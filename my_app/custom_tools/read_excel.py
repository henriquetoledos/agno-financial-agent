#%%
import pandas as pd
import os 
from openpyxl import load_workbook

# Define the path to the Excel file
excel_file_path = os.path.join(os.path.dirname(__file__), '../data/knowledge', 'Extract database.xlsx')

# Read the Excel file
def read_excel_file(file_path=None):
    if os.path.exists(file_path):
        data = pd.read_excel(file_path)
        return data
    else:
        raise FileNotFoundError(f"Excel file not found at {file_path}")


def get_sheet_names():
    if os.path.exists(excel_file_path):
        workbook = load_workbook(excel_file_path)
        return workbook.sheetnames
    else:
        raise FileNotFoundError(f"Excel file not found at {excel_file_path}")
#%%

# Read all of the sheet names
sheet_names = get_sheet_names()

## Read all excel files sheets with pandas
data = pd.read_excel(excel_file_path,sheet_name=sheet_names[0])
#%%

# Drop columns with more than 90% NA values
threshold = 0.9
non_na_threshold = len(data) * (1 - threshold)
filtered_data = data.dropna(axis=1, thresh=non_na_threshold)

# Check for columns with "Unnamed" in their names
unnamed_threshold = 0.8  # Adjust the threshold as needed

while True:
    unnamed_count = sum("Unnamed" in str(col) or pd.isna(col) for col in filtered_data.columns)
    if unnamed_count / len(filtered_data.columns) > unnamed_threshold:
        # Drop the current column names and use the next row as column names
        filtered_data.columns = filtered_data.iloc[0]
        filtered_data = filtered_data[1:].reset_index(drop=True)
    else:
        break

# Drop rows with majority of NA values
min_non_na = int(len(filtered_data.columns) * 0.8)  # Require at least 60% non-NA values

filtered_data.dropna(thresh=min_non_na,inplace=True,axis=0)


