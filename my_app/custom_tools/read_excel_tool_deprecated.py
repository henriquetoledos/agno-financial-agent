import pandas as pd
import os 
from openpyxl import load_workbook

#TODO: Trim and add underscore to name of file and column names 
#TODO:

# Define the path to the Excel file
excel_file_path = os.path.join(os.path.dirname(__file__), '../data/knowledge', 'Extract database.xlsx')

def read_excel_file(file_path=None):
    """Read an Excel file and return the data as a pandas DataFrame."""
    if file_path is None:
        file_path = excel_file_path
        
    if os.path.exists(file_path):
        data = pd.read_excel(file_path)
        return data
    else:
        raise FileNotFoundError(f"Excel file not found at {file_path}")

def get_sheet_names(file_path=None):
    """Get all sheet names from an Excel file."""
    if file_path is None:
        file_path = excel_file_path
        
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        return workbook.sheetnames
    else:
        raise FileNotFoundError(f"Excel file not found at {file_path}")

def process_sheet_data(data):
    """Clean and process the data from a sheet."""
    # Drop columns with more than 90% NA values
    threshold = 0.9
    non_na_threshold = len(data) * (1 - threshold)
    filtered_data = data.dropna(axis=1, thresh=non_na_threshold)

    # Check for columns with "Unnamed" in their names
    unnamed_threshold = 0.8

    while True:
        unnamed_count = sum("Unnamed" in str(col) or pd.isna(col) for col in filtered_data.columns)
        if unnamed_count / len(filtered_data.columns) > unnamed_threshold:
            # Drop the current column names and use the next row as column names
            filtered_data.columns = filtered_data.iloc[0]
            filtered_data = filtered_data[1:].reset_index(drop=True)
        else:
            break

    # Drop rows with majority of NA values
    min_non_na = int(len(filtered_data.columns) * 0.8)
    filtered_data.dropna(thresh=min_non_na, inplace=True, axis=0)
    
    return filtered_data

def excel_to_csv(excel_file=None, output_dir=None):
    """
    Process all sheets in an Excel file and save each as a CSV file.
    
    Args:
        excel_file: Path to the Excel file. If None, uses the default path.
        output_dir: Directory to save CSV files. If None, uses '../data/knowledge/csv'.
    
    Returns:
        list: Paths to all created CSV files
    """
    if excel_file is None:
        excel_file = excel_file_path
    
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(__file__), '../data/knowledge/csv')
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Get filename without extension
    file_basename = os.path.splitext(os.path.basename(excel_file))[0]
    
    # Get all sheet names
    sheet_names = get_sheet_names(excel_file)
    
    csv_files = []
    
    # Process each sheet
    for sheet_name in sheet_names:
        # Read the sheet
        data = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        # Process the data
        processed_data = process_sheet_data(data)
        
        # Create CSV filename
        csv_filename = f"{file_basename}_{sheet_name}.csv"
        csv_path = os.path.join(output_dir, csv_filename)
        
        # Save as CSV
        processed_data.to_csv(csv_path, index=False)
        csv_files.append(csv_path)
        
        print(f"Saved {csv_path}")
    
    return csv_files

# If the script is run directly, process all sheets and save as CSV
if __name__ == "__main__":
    excel_to_csv()
